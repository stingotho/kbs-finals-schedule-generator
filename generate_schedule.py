import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
import argparse

def extract_schedule(teacher_name, proctoring_path, roster_path, output_excel):
    # Load full roster with morning + after-exam duties on same sheet
    roster_df = pd.read_excel(roster_path, header=None)
    
    # Identify the row where after exam duty starts
    after_exam_idx = roster_df[roster_df[0].astype(str).str.contains("After Exams Duty Roster", case=False, na=False)].index[0]
    
    # Extract dates
    dates = roster_df.iloc[2, 1:8].astype(str).tolist()

    # Extract morning and after exam duty tables
    morning_df = roster_df.iloc[3:after_exam_idx, :8].copy()
    after_df = roster_df.iloc[after_exam_idx + 2:, :8].copy()
    morning_df.columns = ["Duty Role"] + dates
    after_df.columns = ["Duty Role"] + dates

    # Extract duties for the target teacher
    final_schedule = []
    for date in dates:
        morning = morning_df.loc[morning_df[date].astype(str).str.contains(teacher_name, na=False), "Duty Role"]
        after = after_df.loc[after_df[date].astype(str).str.contains(teacher_name, na=False), "Duty Role"]
        final_schedule.append({
            "Date": date,
            "Morning Duty": morning.values[0] if not morning.empty else "—",
            "After Exam Duty": after.values[0] if not after.empty else "—"
        })

    # Load proctoring schedule
    proctoring_df = pd.read_excel(proctoring_path)
    proctoring_df.columns = ["Room"] + dates

    # Match proctoring room
    for row in final_schedule:
        date = row["Date"]
        match = proctoring_df[proctoring_df[date].astype(str).str.contains(teacher_name, na=False)]
        row["Proctoring Room"] = match["Room"].values[0] if not match.empty else "—"

    # Create output DataFrame
    result_df = pd.DataFrame(final_schedule)[["Date", "Proctoring Room", "Morning Duty", "After Exam Duty"]]
    result_df.to_excel(output_excel, index=False)

    # Add borders and wrap text
    wb = load_workbook(output_excel)
    ws = wb.active
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(wrap_text=True, vertical='top')
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = align
    wb.save(output_excel)
    print(f"✅ Schedule for {teacher_name} saved to: {output_excel}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate teacher schedule from duty rosters.")
    parser.add_argument("--teacher", required=True, help="Full or partial name of the teacher")
    parser.add_argument("--proctoring", default="Main Proctoring.xlsx", help="Path to proctoring Excel file")
    parser.add_argument("--roster", default="Main Roster Schedule.xlsx", help="Path to duty roster Excel file")
    parser.add_argument("--output", default="Final_Schedule.xlsx", help="Output Excel file name")
    args = parser.parse_args()
    
    extract_schedule(args.teacher, args.proctoring, args.roster, args.output)
