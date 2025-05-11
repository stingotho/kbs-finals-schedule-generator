import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from fpdf import FPDF
import requests
from io import BytesIO

# GitHub-hosted Excel file URLs
GITHUB_RAW = "https://raw.githubusercontent.com/stingotho/kbs-finals-schedule-generator/main/"
DEFAULT_PROCTORING_URL = GITHUB_RAW + "Main%20Proctoring.xlsx"
DEFAULT_ROSTER_URL = GITHUB_RAW + "Main%20Roster%20Schedule.xlsx"

def load_excel_from_url(url):
    response = requests.get(url)
    return BytesIO(response.content)

def extract_teacher_names(proctoring_df, morning_df, after_df, dates):
    names = set()
    for date in dates:
        names.update(proctoring_df[date].dropna().astype(str).values)
        names.update(morning_df[date].dropna().astype(str).values)
        names.update(after_df[date].dropna().astype(str).values)
    flat = set()
    for val in names:
        for part in val.split("+"):
            if part.strip():
                flat.add(part.strip())
    return sorted(flat)

def generate_excel(schedule, output_path):
    result_df = pd.DataFrame(schedule)[["Date", "Proctoring Room", "Morning Duty", "After Exam Duty"]]
    result_df.to_excel(output_path, index=False)
    wb = load_workbook(output_path)
    ws = wb.active
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    align = Alignment(wrap_text=True, vertical='top')
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = align
    wb.save(output_path)
    return result_df

def generate_pdf(schedule, output_path):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)
    pdf.set_fill_color(220, 220, 220)
    col_widths = [45, 45, 50, 50]
    headers = ["Date", "Proctoring Room", "Morning Duty", "After Exam Duty"]
    for i, header in enumerate(headers):
        pdf.cell(col_widths[i], 10, header.encode('latin-1', 'replace').decode('latin-1'), border=1, fill=True)
    pdf.ln()
    for row in schedule:
        pdf.cell(col_widths[0], 10, row["Date"].encode('latin-1', 'replace').decode('latin-1'), border=1)
        pdf.cell(col_widths[1], 10, row["Proctoring Room"].encode('latin-1', 'replace').decode('latin-1'), border=1)
        pdf.cell(col_widths[2], 10, row["Morning Duty"].encode('latin-1', 'replace').decode('latin-1'), border=1)
        pdf.cell(col_widths[3], 10, row["After Exam Duty"].encode('latin-1', 'replace').decode('latin-1'), border=1)
        pdf.ln()
    pdf.output(output_path)

st.set_page_config(page_title="📅 KBS Schedule Generator", layout="centered")
st.title("📅 KBS Finals Schedule Generator")

uploaded_proctor = st.file_uploader("Upload Proctoring Excel File", type="xlsx", key="proctor")
uploaded_roster = st.file_uploader("Upload Duty Roster Excel File", type="xlsx", key="roster")

# Load default if not uploaded
proctoring_file = uploaded_proctor or load_excel_from_url(DEFAULT_PROCTORING_URL)
roster_file = uploaded_roster or load_excel_from_url(DEFAULT_ROSTER_URL)

if proctoring_file and roster_file:
    roster_df = pd.read_excel(roster_file, header=None)
    after_exam_idx = roster_df[roster_df[0].astype(str).str.contains("After Exams Duty Roster", case=False, na=False)].index[0]
    dates = roster_df.iloc[2, 1:8].astype(str).tolist()

    morning_df = roster_df.iloc[3:after_exam_idx, :8].copy()
    after_df = roster_df.iloc[after_exam_idx + 2:, :8].copy()
    morning_df.columns = ["Duty Role"] + dates
    after_df.columns = ["Duty Role"] + dates

    proctoring_df = pd.read_excel(proctoring_file)
    proctoring_df.columns = ["Room"] + dates

    all_teachers = extract_teacher_names(proctoring_df, morning_df, after_df, dates)
    show_all = st.checkbox("Show schedules for all teachers")

    if show_all:
        for selected_teacher in all_teachers:
            st.subheader(f"Schedule: {selected_teacher}")
            schedule = []
            for date in dates:
                morning = morning_df.loc[morning_df[date].astype(str).str.contains(selected_teacher, na=False), "Duty Role"]
                after = after_df.loc[after_df[date].astype(str).str.contains(selected_teacher, na=False), "Duty Role"]
                proctoring = proctoring_df.loc[proctoring_df[date].astype(str).str.contains(selected_teacher, na=False), "Room"]
                schedule.append({
                    "Date": date,
                    "Proctoring Room": proctoring.values[0] if not proctoring.empty else "—",
                    "Morning Duty": morning.values[0] if not morning.empty else "—",
                    "After Exam Duty": after.values[0] if not after.empty else "—"
                })
            st.dataframe(pd.DataFrame(schedule))
    else:
        selected_teacher = st.selectbox("Select a teacher:", all_teachers)
        export_format = st.radio("Choose export format:", ["Excel", "PDF", "Both"])
        schedule = []
        for date in dates:
            morning = morning_df.loc[morning_df[date].astype(str).str.contains(selected_teacher, na=False), "Duty Role"]
            after = after_df.loc[after_df[date].astype(str).str.contains(selected_teacher, na=False), "Duty Role"]
            proctoring = proctoring_df.loc[proctoring_df[date].astype(str).str.contains(selected_teacher, na=False), "Room"]
            schedule.append({
                "Date": date,
                "Proctoring Room": proctoring.values[0] if not proctoring.empty else "—",
                "Morning Duty": morning.values[0] if not morning.empty else "—",
                "After Exam Duty": after.values[0] if not after.empty else "—"
            })

        st.subheader(f"Schedule for: {selected_teacher}")
        result_df = pd.DataFrame(schedule)
        st.dataframe(result_df)

        if export_format in ["Excel", "Both"]:
            excel_path = f"{selected_teacher.replace(' ', '_')}_Schedule.xlsx"
            generate_excel(schedule, excel_path)
            with open(excel_path, "rb") as f:
                st.download_button("Download Excel", f, file_name=excel_path)

        if export_format in ["PDF", "Both"]:
            pdf_path = f"{selected_teacher.replace(' ', '_')}_Schedule.pdf"
            generate_pdf(schedule, pdf_path)
            with open(pdf_path, "rb") as f:
                st.download_button("Download PDF", f, file_name=pdf_path)
